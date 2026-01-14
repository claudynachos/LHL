#!/usr/bin/env python3
"""
Parse list.xlsx to extract player, goalie, and coach data
"""
import openpyxl
import json
import os

def parse_excel_data(filepath):
    """Parse the Excel file and extract all data"""
    workbook = openpyxl.load_workbook(filepath)
    
    data = {
        'players': [],
        'goalies': [],
        'coaches': []
    }
    
    # Parse Players sheet
    if 'Players' in workbook.sheetnames:
        sheet = workbook['Players']
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If name exists
                player = {
                    'name': row[0],
                    'position': row[1] if len(row) > 1 else 'C',
                    'off': int(row[2]) if len(row) > 2 and row[2] else 75,
                    'def': int(row[3]) if len(row) > 3 and row[3] else 75,
                    'phys': int(row[4]) if len(row) > 4 and row[4] else 75,
                    'lead': int(row[5]) if len(row) > 5 and row[5] else 75,
                    'const': int(row[6]) if len(row) > 6 and row[6] else 75,
                }
                data['players'].append(player)
    
    # Parse Goalies sheet
    if 'Goalies' in workbook.sheetnames:
        sheet = workbook['Goalies']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If name exists
                goalie = {
                    'name': row[0],
                    'position': 'G',
                    'off': int(row[1]) if len(row) > 1 and row[1] else 80,
                    'def': int(row[2]) if len(row) > 2 and row[2] else 80,
                    'phys': int(row[3]) if len(row) > 3 and row[3] else 80,
                    'lead': int(row[4]) if len(row) > 4 and row[4] else 75,
                    'const': int(row[5]) if len(row) > 5 and row[5] else 75,
                }
                data['goalies'].append(goalie)
    
    # Parse Coaches sheet
    if 'Coaches' in workbook.sheetnames:
        sheet = workbook['Coaches']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If name exists
                coach = {
                    'name': row[0],
                    'rating': int(row[1]) if len(row) > 1 and row[1] else 75,
                }
                data['coaches'].append(coach)
    
    # If sheets don't exist, try parsing a single sheet with all data
    if not data['players'] and not data['goalies'] and not data['coaches']:
        sheet = workbook.active
        print(f"Using active sheet: {sheet.title}")
        print(f"Columns: {[cell.value for cell in sheet[1]]}")
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If name exists
                # Check if it's a goalie based on position or sheet structure
                position = row[1] if len(row) > 1 and row[1] else 'C'
                
                player_data = {
                    'name': row[0],
                    'position': position,
                    'off': int(row[2]) if len(row) > 2 and row[2] is not None else 75,
                    'def': int(row[3]) if len(row) > 3 and row[3] is not None else 75,
                    'phys': int(row[4]) if len(row) > 4 and row[4] is not None else 75,
                    'lead': int(row[5]) if len(row) > 5 and row[5] is not None else 75,
                    'const': int(row[6]) if len(row) > 6 and row[6] is not None else 75,
                }
                
                if position == 'G':
                    data['goalies'].append(player_data)
                else:
                    data['players'].append(player_data)
    
    return data

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_path = os.path.join(project_root, 'list.xlsx')
    output_path = os.path.join(project_root, 'data', 'parsed_data.json')
    
    print(f"Parsing {excel_path}...")
    data = parse_excel_data(excel_path)
    
    print(f"\nFound:")
    print(f"  Players: {len(data['players'])}")
    print(f"  Goalies: {len(data['goalies'])}")
    print(f"  Coaches: {len(data['coaches'])}")
    
    # Save to JSON
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f"\nData saved to {output_path}")
    
    # Show sample data
    if data['players']:
        print(f"\nSample player: {data['players'][0]}")
    if data['goalies']:
        print(f"Sample goalie: {data['goalies'][0]}")
    if data['coaches']:
        print(f"Sample coach: {data['coaches'][0]}")

if __name__ == '__main__':
    main()
